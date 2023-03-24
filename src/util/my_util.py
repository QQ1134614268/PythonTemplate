# -*- coding:utf-8 -*-
"""
@Time: 2021/8/2
@Description:
"""
import datetime
import random
import re
import string
from base64 import b64encode, b64decode
from binascii import b2a_base64, a2b_base64

from openpyxl.utils import get_column_letter, column_index_from_string
from sqlalchemy import insert

from config.time_conf import DATE_TIME_FORMAT


def s_2_b64(text):
    return b2a_base64(text.encode("utf8"), newline=False).decode("utf8")


def b64_2_s(text):
    # base64转字符串
    return a2b_base64(text.encode("utf8")).decode("utf8")


def s_2_b64_v2(text):
    return b64encode(text.encode("utf8")).decode("utf8")


def b64_2_s_v2(text):
    # base64转字符串
    return b64decode(text.encode("utf8")).decode("utf8")


if __name__ == '__main__':
    text1 = "永恒a"
    print(text1 == b64_2_s(s_2_b64(text1)))
    print(text1 == b64_2_s_v2(s_2_b64_v2(text1)))


class MyStrUti:

    @staticmethod
    def _stripe(lines):
        """
        去除换行,前后空格,MyStrUti
        :param lines:
        :return:
        """
        new_lines = []
        for line in lines:
            line = line.replace("\n", "").replace("\r", "").strip()
            if line:
                new_lines.append(line)
        return new_lines

    @staticmethod
    def first_letter_case(line, upper=True):
        """
        首字母大小写
        :param line:
        :param upper:
        :return:
        """
        arr = list(line)
        if upper:
            arr[0] = arr[0].upper()
        else:
            arr[0] = arr[0].lower()
        line = "".join(arr)
        return line

    @staticmethod
    def underline_case(line, upper=True):
        """
        下滑线大小写
        :param line:
        :param upper:
        :return:
        """
        arr = line.split("_")

        it = iter(arr)
        res_str = next(it)
        for i in it:
            res_str += MyStrUti.first_letter_case(i, upper)
        return res_str

    @staticmethod
    def en(line, upper=True):
        """
        提取英文
        :param line:
        :param upper:
        :return:
        """
        ...

    @staticmethod
    def cn(line, upper=True):
        """
        提取英文
        :param line:
        :param upper:
        :return:
        """
        ...

    @staticmethod
    def num(line, fmt=r"\d+\.?\d*") -> list:
        """
        提取数字
        :param line:
        :param fmt:
        :return:
        """
        return re.findall(line, fmt)

    @staticmethod
    def template_tree(template, obj):
        """
        树形格式化

        :return:

        """

        class Node:
            def __init__(self, data, tpl, res):
                self.data = data
                self.template = tpl
                self.res = res
                self.children = []

            @staticmethod
            def fmt():
                for item in obj["data"]:
                    template.format(**item)
                    obj["res"].append(template)

            @staticmethod
            def add_node(data, tpl, res):
                return Node(data, tpl, res)

            @staticmethod
            def add_str(data, tpl, res):
                return Node(data, tpl, res)

        return

    @staticmethod
    def start_flag(line, upper=True):
        """
        开始标志
        :param line:
        :param upper:
        :return:
        """
        ...


class Data:
    @staticmethod
    def get_data(f_type, length):
        # r"\d+\.?\d*"
        if "int" in f_type:
            return Data.crate_int(length)
        elif "decimal" in f_type:
            return Data.crate_int(length)
        elif "date" in f_type:
            return Data.crate_date()
        else:
            return Data.crate_str(length)

    @staticmethod
    def crate_int(length):
        return random.randint(0, length)

    @staticmethod
    def crate_str(length):
        length = min(random.randint(0, length), 10)
        return ''.join(random.sample(string.ascii_letters + string.digits, length))

    @staticmethod
    def crate_date():
        r_time = datetime.datetime.now() + datetime.timedelta(minutes=random.randint(-1000, 1000))
        return r_time.strftime("%Y-%m-%d")


class SqlUtil:
    @staticmethod
    def insert_data(cla, d_list, u_clo, db_session, step=1000):
        if not u_clo:
            item = d_list[0].items()
            u_clo = {}
            for k, v in item:
                u_clo[k] = getattr(insert(cla).inserted, k)
        for i in range((len(d_list) + 1) // step):
            r_data = d_list[i * step:(i + 1) * step]
            insert_stmt = insert(cla).values(r_data)
            on_duplicate_key_stmt = insert_stmt.on_duplicate_key_update(**u_clo)
            db_session.execute(on_duplicate_key_stmt)
        db_session.commit()


class TimeUtil:
    @staticmethod
    def get_time_str(date_time_format=DATE_TIME_FORMAT):
        return datetime.datetime.now().strftime(date_time_format)


class ExcelUtil:
    # 读取Excel,数据解析
    #     1. 换行
    #     2. 空格
    #     3. 空值
    #     4. bool 混合 大小写 true false
    #
    # 结论: isinstance split() replace('\n','') upper() in ==
    ...

    @staticmethod
    def get_merged_cells(sheet):
        merged_cells = []
        for item in sheet.merged_cells.ranges:
            a, b, c, d = item.bounds
            for i in range(b, d + 1):
                for j in range(a, c + 1):
                    merged_cells.append((i, j))
        return merged_cells

    @staticmethod
    def get_column_letter(num):
        return get_column_letter(num)

    @staticmethod
    def column_index_from_string(txt):
        return column_index_from_string(txt)

    @staticmethod
    def get_data(cell, row_start, row_end, col_start, col_end):
        """
        行数据
        :param cell:
        :param row_start:
        :param row_end:
        :param col_start:
        :param col_end:
        :return:
        """
        data = []
        for col in range(col_start, col_end + 1):
            data_row = []
            for row in range(row_start, row_end + 1):
                data_row.append(cell[row, col])
            data.append(data_row)
        # Util.print_data(data)
        return data
